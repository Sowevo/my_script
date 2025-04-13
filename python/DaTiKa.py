import json

import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils import get_column_letter, column_index_from_string
import os

# 配置常量
CONFIG = {
    "CELL_WIDTH": 2,  # 单元格宽度（字符）
    "CELL_HEIGHT": 12,  # 单元格高度（像素）
    "MAX_ROWS": 38,  # 每列最大行数
    "START_ROW": 9,  # 题目起始行
    "COL_SPACING": 10,  # 题目类型之间的列间距
    "BORDER": Border(  # 单元格边框样式
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    ),
    "HEADER_ALIGNMENT": Alignment(horizontal='center', vertical='center'),  # 标题对齐
    "OPTION_ALIGNMENT": Alignment(horizontal='distributed', vertical='center', justifyLastLine=True)  # 选项对齐
}


def create_answer_sheet(data):
    """创建答题卡"""
    wb = openpyxl.Workbook()
    # 删除默认的工作表
    ws = wb.active
    wb.remove(ws)
    for item in data:
        title = item["title"]
        types = item["types"]
        ws = wb.create_sheet(title=title)
        # 设置页面布局
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        # 设置单元格尺寸
        for col in range(1, 100):
            ws.column_dimensions[get_column_letter(col)].width = CONFIG["CELL_WIDTH"]
        for row in range(1, 100):
            ws.row_dimensions[row].height = CONFIG["CELL_HEIGHT"]
        # 创建标题
        current_col = 2
        col_count = 1

        # 创建各题型区域
        for question_type in types:
            type_col_count = 1
            title_start = f'{get_column_letter(current_col)}6'
            current_row = CONFIG["START_ROW"]
            # 处理每个问题组
            for question in question_type["question"]:
                # 检测行空间是否足够
                required_rows = 2 + (question["to"] - question["from"] + 1)
                if current_row + required_rows > CONFIG["MAX_ROWS"]:
                    # 行空间不足换列的时候,画一下边框
                    apply_thick_border(ws, current_col, current_col + 8, CONFIG["START_ROW"], current_row - 1)
                    type_col_count += 1
                    col_count += 1
                    current_col += CONFIG["COL_SPACING"]
                    current_row = CONFIG["START_ROW"]
                # 创建问题标题
                merge_and_style(ws, f'{get_column_letter(current_col)}{current_row}', 8, 1,
                                question["title"], CONFIG["HEADER_ALIGNMENT"], 18)
                current_row += 2

                # 创建题目选项
                option_num = question.get("option_num", 4)
                for q_num in range(question["from"], question["to"] + 1):
                    # 题号
                    merge_and_style(ws, f'{get_column_letter(current_col)}{current_row}', 1, 0,
                                    q_num, CONFIG["HEADER_ALIGNMENT"], 8, "娃娃体-简", True)
                    # 选项区域
                    option_col = current_col + 2
                    merge_and_style(ws, f'{get_column_letter(option_col)}{current_row}', 6, 0,
                                    '①②③④'[:option_num], CONFIG["OPTION_ALIGNMENT"], 8)
                    current_row += 1
            col_count += 1
            apply_thick_border(ws, current_col, current_col + 8, CONFIG["START_ROW"], current_row - 1)

            # 题型中的最后一列处理完成后再添加题型标题
            # 创建题型标题
            merge_and_style(ws, title_start, type_col_count * 10 - 2, 1,
                            question_type["type"], CONFIG["HEADER_ALIGNMENT"], 18)

            # 切换到下一题型列
            current_col += CONFIG["COL_SPACING"]

        # 最后一列处理完成后再添加总标题
        merge_and_style(ws, f'B1', col_count * 10 - 12, 3, f'{title}', CONFIG["HEADER_ALIGNMENT"], 44)

        print(f"答题卡已全部生成: {title}")

    # 保存文件
    filename = "Answer_Sheet.xlsx"
    wb.save(filename)
    print(f"答题卡已生成: {os.path.abspath(filename)}")


def merge_and_style(ws, start_cell, right_steps, down_steps, content, alignment, size, font_name='等线', bold=False):
    """合并单元格并应用样式"""
    # 设置边框
    start_col, start_row = coordinate_from_string(start_cell)
    start_col_idx = column_index_from_string(start_col)
    for row in range(start_row, start_row + down_steps + 1):
        for col in range(start_col_idx, start_col_idx + right_steps + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = CONFIG["BORDER"]

    # 合并单元格
    end_col = get_column_letter(start_col_idx + right_steps)
    end_row = start_row + down_steps
    ws.merge_cells(f'{start_cell}:{end_col}{end_row}')

    # 设置内容和样式
    cell = ws[start_cell]
    cell.value = content
    cell.alignment = alignment
    # 文字大小
    cell.font = Font(name=font_name, size=size, bold=bold)
    return cell


def apply_thick_border(ws, start_col, end_col, start_row, end_row):
    """在指定区域应用粗边框"""
    thick_side = Side(style='medium')

    def get_updated_border(original, **kwargs):
        """创建新边框对象，保留未修改的属性"""
        return Border(
            left=kwargs.get('left', original.left),
            right=kwargs.get('right', original.right),
            top=kwargs.get('top', original.top),
            bottom=kwargs.get('bottom', original.bottom),
            diagonal=kwargs.get('diagonal', original.diagonal),
            diagonal_direction=kwargs.get('diagonal_direction', original.diagonal_direction),
            vertical=kwargs.get('vertical', original.vertical),
            horizontal=kwargs.get('horizontal', original.horizontal)
        )

    # 设置左边框
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=start_col)
        new_border = get_updated_border(cell.border, left=thick_side)
        cell.border = new_border

    # 设置右边框
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=end_col)
        new_border = get_updated_border(cell.border, right=thick_side)
        cell.border = new_border

    # 设置顶部边框
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=start_row, column=col)
        new_border = get_updated_border(cell.border, top=thick_side)
        cell.border = new_border

    # 设置底部边框
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=end_row, column=col)
        new_border = get_updated_border(cell.border, bottom=thick_side)
        cell.border = new_border


# 运行主函数
json_data = json.load(open("DaTiKa.json", encoding="utf-8"))
create_answer_sheet(json_data)
